from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config.settings import (
    EXPORT_DRY_RUN,
    EXPORT_INCLUDE_REJECTED,
    EXPORT_OUTPUT_DIR,
)
from database_manager import Database

from .models import ExportDataset, ExportDryRunResult, ExportResult, ExportSalon
from .query_service import ExportQueryService


ACCEPTED_HEADERS = [
    "Регион",
    "Город / населённый пункт",
    "Название",
    "Адрес",
    "Телефон",
    "Сайт / соцсети",
    "Цена базового маникюра с покрытием",
    "Валюта",
    "Тип цены",
    "Название услуги",
    "Источник цены",
    "Статус цены",
    "Количество мастеров",
    "Тип салона",
    "Статус проверки",
    "Комментарий",
    "Источник организации",
    "External ID",
    "Дата обнаружения",
    "Дата последней проверки",
    "Версия классификатора",
]

EXCLUDED_HEADERS = [
    "Регион",
    "Название",
    "Адрес",
    "Причина исключения",
    "Профиль бизнеса",
    "Коды решения",
    "External ID",
    "Версия классификатора",
    "Дата классификации",
]


class ExcelExporter:
    """Render accepted salon export workbooks."""

    HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
    HEADER_FONT = Font(bold=True)
    TOP_WRAP = Alignment(vertical="top", wrap_text=True)

    def __init__(
        self,
        database: Database,
        query_service: ExportQueryService | None = None,
        output_dir: str = EXPORT_OUTPUT_DIR,
        include_rejected: bool = EXPORT_INCLUDE_REJECTED,
        dry_run: bool = EXPORT_DRY_RUN,
    ) -> None:
        self.database = database
        self.query_service = query_service or ExportQueryService(database)
        self.output_dir = Path(output_dir)
        self.include_rejected = include_rejected
        self.dry_run = dry_run

    def export(
        self,
        output_path: str | Path | None = None,
        region_id: int | None = None,
    ) -> ExportResult | ExportDryRunResult:
        """Create an Excel workbook atomically or return a dry-run preview."""

        dataset = self.query_service.load_dataset(region_id=region_id)
        final_path = Path(output_path) if output_path else self.default_output_path()

        if self.dry_run:
            return ExportDryRunResult(
                output_path=str(final_path),
                accepted_count=len(dataset.accepted_salons),
                rejected_count=len(dataset.excluded_salons),
            )

        final_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = self._build_workbook(dataset)
        fd, temporary_name = tempfile.mkstemp(
            prefix=f".{final_path.name}.",
            suffix=".tmp",
            dir=final_path.parent,
        )
        os.close(fd)
        temporary_path = Path(temporary_name)

        try:
            workbook.save(temporary_path)
            os.replace(temporary_path, final_path)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

        return ExportResult(
            output_path=str(final_path),
            accepted_count=len(dataset.accepted_salons),
            rejected_count=len(dataset.excluded_salons),
            file_size_bytes=final_path.stat().st_size,
        )

    def default_output_path(self) -> Path:
        """Return the timestamped default workbook path."""

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        return self.output_dir / f"beauty_salons_{timestamp}.xlsx"

    def _build_workbook(self, dataset: ExportDataset) -> Workbook:
        workbook = Workbook()
        salons_sheet = workbook.active
        salons_sheet.title = "Салоны"
        self._write_accepted_sheet(salons_sheet, dataset.accepted_salons)
        report_sheet = workbook.create_sheet("Отчёт")
        self._write_report_sheet(report_sheet, dataset)

        if self.include_rejected:
            excluded_sheet = workbook.create_sheet("Исключённые")
            self._write_excluded_sheet(excluded_sheet, dataset)

        return workbook

    def _write_accepted_sheet(
        self,
        sheet: Worksheet,
        salons: list[ExportSalon],
    ) -> None:
        self._append_header(sheet, ACCEPTED_HEADERS)

        for salon in salons:
            price = salon.price
            sheet.append(
                [
                    salon.region,
                    salon.city,
                    salon.name,
                    salon.address,
                    "\n".join(salon.phones),
                    "\n".join(salon.links),
                    price.display_value if price else None,
                    price.currency if price else None,
                    price.price_type if price else "not_checked",
                    price.service_name if price else None,
                    price.source_type if price else None,
                    price.status if price else "not_checked",
                    salon.masters_count,
                    self._business_profile_label(salon.business_profile),
                    salon.verification_status,
                    salon.comment,
                    salon.external_source,
                    salon.external_id,
                    salon.first_seen_at,
                    salon.last_checked_at,
                    salon.classifier_version,
                ]
            )

        self._format_table(sheet, ACCEPTED_HEADERS)
        self._format_date_columns(sheet, (19, 20))

    def _write_excluded_sheet(
        self,
        sheet: Worksheet,
        dataset: ExportDataset,
    ) -> None:
        self._append_header(sheet, EXCLUDED_HEADERS)

        for salon in dataset.excluded_salons:
            sheet.append(
                [
                    salon.region,
                    salon.name,
                    salon.address,
                    salon.rejection_reason,
                    salon.business_profile,
                    salon.reason_codes,
                    salon.external_id,
                    salon.classifier_version,
                    salon.classified_at,
                ]
            )

        self._format_table(sheet, EXCLUDED_HEADERS)
        self._format_date_columns(sheet, (9,))

    def _write_report_sheet(
        self,
        sheet: Worksheet,
        dataset: ExportDataset,
    ) -> None:
        accepted = dataset.accepted_salons
        excluded = dataset.excluded_salons
        with_contacts = sum(1 for salon in accepted if salon.phones or salon.links)
        with_links = sum(1 for salon in accepted if salon.links)
        with_prices = sum(
            1
            for salon in accepted
            if salon.price and salon.price.price_type in ("exact", "from", "range")
        )
        without_prices = len(accepted) - with_prices

        rows = [
            ("export_timestamp", datetime.now()),
            ("database_path", dataset.database_path),
            ("total_regions", dataset.total_regions),
            ("regions_with_generated_grids", dataset.regions_with_grids),
            ("accepted_salon_count", len(accepted)),
            ("rejected_salon_count", len(excluded)),
            ("accepted_salons_with_phones", sum(1 for salon in accepted if salon.phones)),
            ("accepted_salons_with_websites_socials", with_links),
            ("accepted_salons_with_active_target_price", with_prices),
            ("accepted_salons_without_contacts", len(accepted) - with_contacts),
            ("accepted_salons_without_prices", without_prices),
            ("current_classifier_version", dataset.classifier_version),
            ("workbook_generation_status", "complete"),
        ]
        sheet.append(["Метрика", "Значение"])

        for row in rows:
            sheet.append(list(row))

        sheet.append([])
        sheet.append(["Регион", "Accepted", "Rejected", "With contacts", "With prices", "Last scan status"])

        for summary in dataset.region_summaries:
            sheet.append(
                [
                    summary.region,
                    summary.accepted,
                    summary.rejected,
                    summary.with_contacts,
                    summary.with_prices,
                    summary.last_scan_status,
                ]
            )

        self._format_table(sheet, ["Метрика", "Значение"])
        self._format_date_columns(sheet, (2,))
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 42

    def _append_header(self, sheet: Worksheet, headers: list[str]) -> None:
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.TOP_WRAP

    def _format_table(self, sheet: Worksheet, headers: list[str]) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        widths = self._column_widths(headers)

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = self.TOP_WRAP

        for row_index in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_index].height = 36

    def _format_date_columns(self, sheet: Worksheet, columns: tuple[int, ...]) -> None:
        for column in columns:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=column)

                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"

    def _column_widths(self, headers: list[str]) -> list[int]:
        default = {
            "Регион": 26,
            "Город / населённый пункт": 18,
            "Название": 32,
            "Адрес": 32,
            "Телефон": 24,
            "Сайт / соцсети": 32,
            "Цена базового маникюра с покрытием": 18,
            "Комментарий": 48,
            "External ID": 22,
            "Коды решения": 42,
        }
        return [default.get(header, 20) for header in headers]

    def _business_profile_label(self, business_profile: str | None) -> str:
        mapping = {
            "nail_specialist": "Маникюрная студия",
            "mixed_beauty_salon": "Смешанный салон красоты",
            "unknown": "Не определён",
            None: "Не определён",
        }
        return mapping.get(business_profile, "Не определён")
