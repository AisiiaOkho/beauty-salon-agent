from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from database_manager import Database
from exporting.excel_exporter import ExcelExporter
from geometry.models import GridCell
from maintenance.canonical_backfill import CanonicalOrganizationBackfiller
from scanner.models import ClassificationResult, RawOrganization, SearchPage
from scanner.salon_scanner import SalonScannerManager


class CanonicalBackfillDatabaseMixin:
    def make_database(self, directory: str) -> Database:
        database = Database(db_path=Path(directory) / "test.db")
        database.create_tables()
        database.sync_regions()
        database.insert_grid_cell_batch(
            1,
            [
                GridCell(
                    cell_order=1,
                    north=54.72,
                    south=54.70,
                    west=20.44,
                    east=20.46,
                    center_lat=54.71,
                    center_lon=20.45,
                )
            ],
        )
        return database

    def insert_discovery(
        self,
        database: Database,
        *,
        external_id: str,
        name: str,
        status: str = "rejected",
        categories: list[str] | None = None,
        reason_codes: list[str] | None = None,
        rejection_reason: str | None = "mixed_non_salon",
        business_profile: str = "mixed_non_salon",
        query: str = "маникюр",
    ) -> None:
        payload = {
            "id": external_id,
            "name": name,
            "address_name": "Ленина, 1",
            "point": {"lat": 54.71, "lon": 20.45},
            "rubrics": [{"name": category} for category in (categories or [])],
        }
        codes = reason_codes or ["manicure_signal", "unrelated_primary_category_signal"]

        with database.connect() as connection:
            cursor = connection.execute(
                """
                INSERT INTO raw_organization_results (
                    region_id,
                    grid_cell_id,
                    query,
                    external_source,
                    external_id,
                    name,
                    payload
                )
                VALUES (1, 1, ?, '2GIS', ?, ?, ?)
                """,
                (query, external_id, name, json.dumps(payload, ensure_ascii=False)),
            )
            raw_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO salon_discoveries (
                    raw_result_id,
                    region_id,
                    grid_cell_id,
                    query,
                    external_source,
                    external_id,
                    filter_status,
                    filter_confidence,
                    filter_reasons,
                    classifier_reason_codes,
                    rejection_reason,
                    business_profile
                )
                VALUES (?, 1, 1, ?, '2GIS', ?, ?, 0.95, ?, ?, ?, ?)
                """,
                (
                    raw_id,
                    query,
                    external_id,
                    status,
                    json.dumps(codes, ensure_ascii=False),
                    json.dumps(codes, ensure_ascii=False),
                    rejection_reason if status == "rejected" else None,
                    business_profile,
                ),
            )
            connection.commit()

    def count_salons(self, database: Database, status: str | None = None) -> int:
        with database.connect() as connection:
            if status is None:
                row = connection.execute("SELECT COUNT(*) AS total FROM salons").fetchone()
            else:
                row = connection.execute(
                    "SELECT COUNT(*) AS total FROM salons WHERE filter_status = ?",
                    (status,),
                ).fetchone()

        return int(row["total"])


class StaticSearchClient:
    def __init__(self, organization: RawOrganization) -> None:
        self.organization = organization

    def search(self, **kwargs: object) -> SearchPage:
        return SearchPage(
            organizations=[
                RawOrganization(
                    external_source=self.organization.external_source,
                    external_id=self.organization.external_id,
                    name=self.organization.name,
                    address=self.organization.address,
                    latitude=self.organization.latitude,
                    longitude=self.organization.longitude,
                    categories=list(self.organization.categories),
                    raw_payload=dict(self.organization.raw_payload),
                    discovered_query=str(kwargs["query"]),
                    discovered_grid_cell_id=int(kwargs["grid_cell_id"]),
                )
            ],
            page=1,
            has_next_page=False,
        )


class RejectingClassifier:
    def classify(self, organization: RawOrganization) -> ClassificationResult:
        del organization
        return ClassificationResult(
            accepted=False,
            confidence=0.95,
            reason_codes=["manicure_signal", "unrelated_primary_category_signal"],
            business_profile="mixed_non_salon",
            rejection_reason="mixed_non_salon",
        )


class CanonicalBackfillTests(CanonicalBackfillDatabaseMixin, unittest.TestCase):
    def test_newly_rejected_discovery_becomes_canonical_rejected_entity(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(
                database,
                external_id="xfit",
                name="Xfit, фитнес-клуб",
                categories=["Ногтевые студии", "Фитнес-клубы"],
            )

            summary = CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()

            self.assertEqual(summary.created, 1)
            self.assertEqual(self.count_salons(database, "rejected"), 1)
            with database.connect() as connection:
                row = connection.execute(
                    "SELECT salon_id FROM salon_discoveries WHERE external_id = 'xfit'"
                ).fetchone()
            self.assertIsNotNone(row["salon_id"])

    def test_repeated_rejected_observations_deduplicated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")

            backfiller = CanonicalOrganizationBackfiller(database, dry_run=False)
            first = backfiller.backfill_rejected()
            second = backfiller.backfill_rejected()

            self.assertEqual(first.created, 1)
            self.assertEqual(second.created, 0)
            self.assertEqual(second.unchanged, 1)
            self.assertEqual(self.count_salons(database), 1)

    def test_accepted_entity_not_duplicated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            organization = RawOrganization(
                external_source="2GIS",
                external_id="same",
                name="K-Studio, студия маникюра",
                categories=["Ногтевые студии"],
                raw_payload={"id": "same", "name": "K-Studio, студия маникюра"},
                discovered_grid_cell_id=1,
            )
            database.upsert_salon(
                1,
                organization,
                ClassificationResult(
                    accepted=True,
                    confidence=1.0,
                    reason_codes=["manicure_signal"],
                    business_profile="nail_specialist",
                ),
            )
            self.insert_discovery(database, external_id="same", name="K-Studio, студия маникюра")

            summary = CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()

            self.assertEqual(summary.created, 0)
            self.assertEqual(summary.updated, 1)
            self.assertEqual(self.count_salons(database), 1)
            self.assertEqual(self.count_salons(database, "rejected"), 1)

    def test_rejected_entity_excluded_from_enrichment_and_pricing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")
            CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()
            with database.connect() as connection:
                salon_id = int(connection.execute("SELECT id FROM salons").fetchone()["id"])

            self.assertIsNone(database.get_salon_for_enrichment_by_id(salon_id))
            self.assertIsNone(database.get_salon_for_pricing_by_id(salon_id))

    def test_rejected_entity_absent_from_accepted_export_and_present_in_excluded(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")
            CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()
            path = Path(directory) / "export.xlsx"

            ExcelExporter(database, dry_run=False).export(output_path=path)
            workbook = load_workbook(path)

            self.assertEqual(workbook["Салоны"].max_row, 1)
            self.assertEqual(workbook["Исключённые"].max_row, 2)
            self.assertEqual(workbook["Исключённые"].cell(2, 2).value, "Xfit, фитнес-клуб")

    def test_accepted_to_rejected_reclassification(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            organization = RawOrganization(
                external_source="2GIS",
                external_id="xfit",
                name="Xfit, фитнес-клуб",
                categories=["Ногтевые студии"],
                raw_payload={"id": "xfit", "name": "Xfit, фитнес-клуб"},
                discovered_grid_cell_id=1,
            )
            salon_id, _ = database.upsert_salon(
                1,
                organization,
                ClassificationResult(
                    accepted=True,
                    confidence=1.0,
                    reason_codes=["manicure_signal"],
                    business_profile="mixed_beauty_salon",
                ),
            )

            database.save_salon_classification_result(
                salon_id,
                ClassificationResult(
                    accepted=False,
                    confidence=0.95,
                    reason_codes=["manicure_signal", "unrelated_primary_category_signal"],
                    business_profile="mixed_non_salon",
                    rejection_reason="mixed_non_salon",
                ),
                input_snapshot={"test": True},
            )

            self.assertIsNone(database.get_salon_for_enrichment_by_id(salon_id))
            self.assertEqual(self.count_salons(database, "rejected"), 1)

    def test_rejected_to_accepted_reclassification(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="salon", name="Beauty, студия красоты")
            CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()
            with database.connect() as connection:
                salon_id = int(connection.execute("SELECT id FROM salons").fetchone()["id"])

            database.save_salon_classification_result(
                salon_id,
                ClassificationResult(
                    accepted=True,
                    confidence=1.0,
                    reason_codes=["manicure_signal", "salon_or_studio_name_signal"],
                    business_profile="mixed_beauty_salon",
                ),
                input_snapshot={"test": True},
            )

            self.assertIsNotNone(database.get_salon_for_enrichment_by_id(salon_id))
            self.assertEqual(self.count_salons(database, "accepted"), 1)

    def test_dry_run_backfill_changes_nothing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")

            summary = CanonicalOrganizationBackfiller(database, dry_run=True).backfill_rejected()

            self.assertEqual(summary.created, 1)
            self.assertEqual(self.count_salons(database), 0)

    def test_legacy_xfit_style_row_compatibility(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(database, external_id="xfit", name="Xfit, фитнес-клуб")
            database.upsert_salon(
                1,
                RawOrganization(
                    external_source="2GIS",
                    external_id="xfit",
                    name="Xfit, фитнес-клуб",
                    categories=["Ногтевые студии", "Фитнес-клубы"],
                    raw_payload={"id": "xfit", "name": "Xfit, фитнес-клуб"},
                ),
                ClassificationResult(
                    accepted=False,
                    confidence=0.95,
                    reason_codes=["manicure_signal", "unrelated_primary_category_signal"],
                    business_profile="mixed_non_salon",
                    rejection_reason="mixed_non_salon",
                ),
            )

            summary = CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()

            self.assertEqual(summary.unchanged, 1)
            self.assertEqual(self.count_salons(database), 1)

    def test_conflicting_decisions_resolved_by_latest_discovery(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_discovery(
                database,
                external_id="same",
                name="Beauty, студия красоты",
                status="accepted",
                categories=["Ногтевые студии"],
                rejection_reason=None,
                business_profile="mixed_beauty_salon",
            )
            self.insert_discovery(
                database,
                external_id="same",
                name="Beauty, студия красоты",
                status="rejected",
            )

            summary = CanonicalOrganizationBackfiller(database, dry_run=False).backfill_rejected()

            self.assertEqual(summary.conflicts, 1)
            self.assertEqual(self.count_salons(database, "rejected"), 1)

    def test_scanner_creates_canonical_rejected_entity(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            organization = RawOrganization(
                external_source="2GIS",
                external_id="xfit",
                name="Xfit, фитнес-клуб",
                categories=["Ногтевые студии", "Фитнес-клубы"],
                raw_payload={"id": "xfit", "name": "Xfit, фитнес-клуб"},
            )
            scanner = SalonScannerManager(
                database,
                search_client=StaticSearchClient(organization),
                classifier=RejectingClassifier(),
                max_cells_per_run=1,
                dry_run=False,
                progress_logger=lambda message: None,
            )

            summary = scanner.scan_region(database.get_region_progress(1))

            self.assertEqual(summary.rejected_results, 5)
            self.assertEqual(self.count_salons(database, "rejected"), 1)


if __name__ == "__main__":
    unittest.main()
