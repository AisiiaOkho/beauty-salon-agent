from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from database_manager import Database
from exporting.excel_exporter import ACCEPTED_HEADERS, ExcelExporter
from geometry.models import GridCell


class ExportDatabaseMixin:
    def make_database(self, directory: str) -> Database:
        database = Database(db_path=Path(directory) / "test.db")
        database.create_tables()
        database.sync_regions()
        return database

    def insert_salon(
        self,
        database: Database,
        *,
        region_id: int = 1,
        name: str = "Студия маникюра",
        external_id: str = "branch-1",
        filter_status: str = "accepted",
        business_profile: str | None = "nail_specialist",
        rejection_reason: str | None = None,
        classifier_version: str | None = "2.0.0",
    ) -> int:
        with database.connect() as connection:
            cursor = connection.execute(
                """
                INSERT INTO salons (
                    region_id,
                    external_source,
                    source,
                    external_id,
                    name,
                    address,
                    filter_status,
                    business_profile,
                    rejection_reason,
                    classifier_reason_codes,
                    classifier_version,
                    classified_at,
                    first_seen_at,
                    updated_at,
                    verification_status
                )
                VALUES (?, '2GIS', '2GIS', ?, ?, 'Ленина, 1', ?, ?, ?, ?, ?, ?, ?, ?, 'not_checked')
                """,
                (
                    region_id,
                    external_id,
                    name,
                    filter_status,
                    business_profile,
                    rejection_reason,
                    '["test"]',
                    classifier_version,
                    "2026-07-24 10:00:00",
                    "2026-07-24 09:00:00",
                    "2026-07-24 11:00:00",
                ),
            )
            connection.commit()

        return int(cursor.lastrowid)

    def add_contact(
        self,
        database: Database,
        salon_id: int,
        contact_type: str,
        value: str,
        active: int = 1,
    ) -> None:
        with database.connect() as connection:
            connection.execute(
                """
                INSERT OR IGNORE INTO salon_contacts (
                    salon_id,
                    contact_type,
                    display_value,
                    normalized_value,
                    source,
                    is_active
                )
                VALUES (?, ?, ?, ?, '2GIS', ?)
                """,
                (salon_id, contact_type, value, value, active),
            )
            connection.commit()

    def add_price(
        self,
        database: Database,
        salon_id: int,
        *,
        price_type: str,
        amount_minor: int | None = None,
        range_min_minor: int | None = None,
        range_max_minor: int | None = None,
        active: int = 1,
    ) -> None:
        with database.connect() as connection:
            connection.execute(
                """
                INSERT INTO salon_prices (
                    salon_id,
                    service_key,
                    service_name_raw,
                    service_name_normalized,
                    amount_minor,
                    currency,
                    price_type,
                    range_min_minor,
                    range_max_minor,
                    source_type,
                    source_url,
                    evidence_text,
                    confidence,
                    is_active
                )
                VALUES (?, 'basic_manicure_with_coating', 'Маникюр с покрытием',
                        'маникюр с покрытием', ?, 'RUB', ?, ?, ?, 'website_text',
                        'https://example.test/prices', ?, 'high', ?)
                """,
                (
                    salon_id,
                    amount_minor,
                    price_type,
                    range_min_minor,
                    range_max_minor,
                    f"Маникюр с покрытием {amount_minor or range_min_minor}",
                    active,
                ),
            )
            connection.commit()


class ExcelExportTests(ExportDatabaseMixin, unittest.TestCase):
    def workbook_path(self, directory: str) -> Path:
        return Path(directory) / "export.xlsx"

    def test_accepted_salon_exported_and_rejected_absent(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database, name="Accepted", external_id="a")
            self.insert_salon(
                database,
                name="Xfit, фитнес-клуб",
                external_id="x",
                filter_status="rejected",
                business_profile="mixed_non_salon",
                rejection_reason="mixed_non_salon",
            )
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            workbook = load_workbook(path)
            names = [row[2].value for row in workbook["Салоны"].iter_rows(min_row=2)]

            self.assertEqual(names, ["Accepted"])
            self.assertNotIn("Xfit, фитнес-клуб", names)

    def test_rejected_salon_present_in_excluded_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(
                database,
                name="Xfit, фитнес-клуб",
                external_id="x",
                filter_status="rejected",
                business_profile="mixed_non_salon",
                rejection_reason="mixed_non_salon",
            )
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Исключённые"]

            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=2).value, "Xfit, фитнес-клуб")

    def test_multiple_contacts_aggregate_without_duplicate_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            salon_id = self.insert_salon(database, name="Contacts", external_id="c")
            self.add_contact(database, salon_id, "phone", "+7 111")
            self.add_contact(database, salon_id, "phone", "+7 111")
            self.add_contact(database, salon_id, "website", "https://example.test")
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=5).value, "+7 111")
            self.assertEqual(sheet.cell(row=2, column=6).value, "https://example.test")

    def test_no_contacts_exports_blank_cells(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database)
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertIsNone(sheet.cell(row=2, column=5).value)
            self.assertIsNone(sheet.cell(row=2, column=6).value)

    def test_price_display_exact_from_range_and_ambiguous(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            exact = self.insert_salon(database, name="Exact", external_id="exact")
            from_price = self.insert_salon(database, name="From", external_id="from")
            range_price = self.insert_salon(database, name="Range", external_id="range")
            ambiguous = self.insert_salon(database, name="Ambiguous", external_id="amb")
            self.add_price(database, exact, price_type="exact", amount_minor=180000)
            self.add_price(database, from_price, price_type="from", amount_minor=150000)
            self.add_price(
                database,
                range_price,
                price_type="range",
                range_min_minor=150000,
                range_max_minor=190000,
            )
            self.add_price(database, ambiguous, price_type="ambiguous")
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]
            values = {sheet.cell(row=i, column=3).value: sheet.cell(row=i, column=7).value for i in range(2, 6)}
            statuses = {sheet.cell(row=i, column=3).value: sheet.cell(row=i, column=12).value for i in range(2, 6)}

            self.assertEqual(values["Exact"], "1800")
            self.assertEqual(values["From"], "от 1500")
            self.assertEqual(values["Range"], "1500–1900")
            self.assertIsNone(values["Ambiguous"])
            self.assertEqual(statuses["Ambiguous"], "ambiguous")

    def test_not_checked_price_status(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database)
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertIsNone(sheet.cell(row=2, column=7).value)
            self.assertEqual(sheet.cell(row=2, column=12).value, "not_checked")

    def test_inactive_contacts_and_prices_excluded(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            salon_id = self.insert_salon(database)
            self.add_contact(database, salon_id, "phone", "+7 inactive", active=0)
            self.add_price(database, salon_id, price_type="exact", amount_minor=180000, active=0)
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertIsNone(sheet.cell(row=2, column=5).value)
            self.assertIsNone(sheet.cell(row=2, column=7).value)

    def test_workbook_sheets_headers_freeze_and_autofilter(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database)
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            workbook = load_workbook(path)
            sheet = workbook["Салоны"]

            self.assertEqual(workbook.sheetnames, ["Салоны", "Отчёт", "Исключённые"])
            self.assertEqual([cell.value for cell in sheet[1]], ACCEPTED_HEADERS)
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertIsNotNone(sheet.auto_filter.ref)

    def test_date_cell_types(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database)
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertIsInstance(sheet.cell(row=2, column=19).value, datetime)
            self.assertIsInstance(sheet.cell(row=2, column=20).value, datetime)

    def test_safe_filename_default(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            exporter = ExcelExporter(database, output_dir=directory)

            path = exporter.default_output_path()

            self.assertEqual(path.suffix, ".xlsx")
            self.assertTrue(path.name.startswith("beauty_salons_"))

    def test_atomic_replacement(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database, name="First", external_id="first")
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            first_size = path.stat().st_size
            self.insert_salon(database, name="Second", external_id="second")
            ExcelExporter(database, dry_run=False).export(output_path=path)

            self.assertTrue(path.exists())
            self.assertNotEqual(path.stat().st_size, 0)
            self.assertNotEqual(path.stat().st_size, first_size)

    def test_region_filter(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database, region_id=1, name="Region one", external_id="one")
            self.insert_salon(database, region_id=2, name="Region two", external_id="two")
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path, region_id=2)
            sheet = load_workbook(path)["Салоны"]

            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=3).value, "Region two")

    def test_dry_run_creates_no_file(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(database)
            path = self.workbook_path(directory)

            result = ExcelExporter(database, dry_run=True).export(output_path=path)

            self.assertFalse(path.exists())
            self.assertFalse(result.file_written)

    def test_many_discoveries_still_one_row(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            salon_id = self.insert_salon(database)
            database.insert_grid_cell_batch(
                1,
                [
                    GridCell(
                        cell_order=1,
                        north=54.72,
                        south=54.70,
                        west=20.44,
                        east=20.46,
                        center_lat=54.71,
                        center_lon=20.45,
                    )
                ],
            )
            with database.connect() as connection:
                for index in range(3):
                    connection.execute(
                        """
                        INSERT INTO salon_discoveries (
                            salon_id,
                            region_id,
                            grid_cell_id,
                            query,
                            external_source,
                            external_id,
                            filter_status,
                            filter_confidence,
                            filter_reasons
                        )
                        VALUES (?, 1, 1, ?, '2GIS', 'branch-1', 'accepted', 1.0, '[]')
                        """,
                        (salon_id, f"query-{index}"),
                    )
                connection.commit()
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertEqual(sheet.max_row, 2)

    def test_legacy_null_classifier_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = self.make_database(directory)
            self.insert_salon(
                database,
                name="Legacy",
                external_id="legacy",
                business_profile=None,
                classifier_version=None,
            )
            path = self.workbook_path(directory)

            ExcelExporter(database, dry_run=False).export(output_path=path)
            sheet = load_workbook(path)["Салоны"]

            self.assertEqual(sheet.cell(row=2, column=14).value, "Не определён")
            self.assertIsNone(sheet.cell(row=2, column=21).value)


if __name__ == "__main__":
    unittest.main()
